"""Excel export module to compile SQL query results into a structured lab report workbook."""

import logging
from decimal import Decimal
import datetime
from pathlib import Path
from typing import Any, List, Tuple
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# Set up logging for this module
logger = logging.getLogger(__name__)


def create_styled_workbook() -> openpyxl.Workbook:
    """Creates a new workbook and removes the default sheet.

    Returns:
        openpyxl.Workbook: The empty, styled workbook.
    """
    wb = openpyxl.Workbook()
    # Remove the default sheet created by openpyxl
    wb.remove(wb.active)
    return wb


def export_query_results_to_sheet(
    wb: openpyxl.Workbook,
    sheet_name: str,
    status: str,
    title: str,
    description: str,
    sql_query: str,
    headers: List[str],
    rows: List[Tuple[Any, ...]],
    error_message: str,
    duration_ms: float,
    timestamp: str,
    database: str
) -> None:
    """Writes query results to a worksheet following a structured report layout.

    Args:
        wb: openpyxl Workbook instance.
        sheet_name: Name of the sheet (e.g. 'Q1').
        status: Run status ('completed' or 'failed').
        title: Exercise title.
        description: Problem statement.
        sql_query: SQL source query string.
        headers: List of column header names.
        rows: List of query result records.
        error_message: Database error message if failed.
        duration_ms: Execution duration in milliseconds.
        timestamp: Execution timestamp.
        database: Connected database name.
    """
    logger.info(
        f"Writing structured worksheet '{sheet_name}' (Status: {status})..."
    )
    ws = wb.create_sheet(title=sheet_name)

    # 1. Fonts and border sides
    title_label_font = Font(name="Segoe UI", size=9, italic=True, color="555555")
    question_font = Font(name="Segoe UI", size=16, bold=True, color="1F497D")
    title_font = Font(name="Segoe UI", size=12, bold=True, color="1F497D")
    section_header_font = Font(name="Segoe UI", size=11, bold=True, color="222222")
    body_font = Font(name="Segoe UI", size=10)
    monospace_font = Font(name="Courier New", size=10, color="000000")
    bold_body_font = Font(name="Segoe UI", size=10, bold=True)
    table_header_font = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
    table_data_font = Font(name="Segoe UI", size=10)
    error_msg_font = Font(name="Courier New", size=10, color="C00000")
    
    thin_gray_side = Side(style="thin", color="D3D3D3")
    thin_gray = Border(
        left=thin_gray_side,
        right=thin_gray_side,
        top=thin_gray_side,
        bottom=thin_gray_side,
    )

    # Row 2: Question Number Header
    ws["A2"] = "Question Number:"
    ws["A2"].font = title_label_font
    ws["A3"] = sheet_name
    ws["A3"].font = question_font

    # Row 5: Exercise Title
    ws["A5"] = "Exercise Title:"
    ws["A5"].font = title_label_font
    ws["A6"] = title
    ws["A6"].font = title_font

    # Row 8: Problem Statement
    ws["A8"] = "Problem Statement:"
    ws["A8"].font = section_header_font

    # Merge A9:H10 for Description Box with light blue fill
    ws.merge_cells("A9:H10")
    ws["A9"] = description
    ws["A9"].font = body_font
    ws["A9"].alignment = Alignment(wrap_text=True, vertical="top")
    info_fill = PatternFill(start_color="F2F5F8", end_color="F2F5F8", fill_type="solid")
    
    for r in range(9, 11):
        for c in range(1, 9):
            cell = ws.cell(row=r, column=c)
            cell.fill = info_fill
            left = thin_gray_side if c == 1 else None
            right = thin_gray_side if c == 8 else None
            top = thin_gray_side if r == 9 else None
            bottom = thin_gray_side if r == 10 else None
            cell.border = Border(left=left, right=right, top=top, bottom=bottom)

    # Row 12: SQL Query Used
    ws["A12"] = "SQL Query Used:"
    ws["A12"].font = section_header_font

    # Merge A13:H16 for SQL Query Box with light gray fill
    ws.merge_cells("A13:H16")
    ws["A13"] = sql_query.strip()
    ws["A13"].font = monospace_font
    ws["A13"].alignment = Alignment(wrap_text=True, vertical="top")
    code_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
    
    for r in range(13, 17):
        for c in range(1, 9):
            cell = ws.cell(row=r, column=c)
            cell.fill = code_fill
            left = thin_gray_side if c == 1 else None
            right = thin_gray_side if c == 8 else None
            top = thin_gray_side if r == 13 else None
            bottom = thin_gray_side if r == 16 else None
            cell.border = Border(left=left, right=right, top=top, bottom=bottom)

    # Row 18: Telemetry Metadata Section
    ws["A18"] = "Execution Metadata:"
    ws["A18"].font = section_header_font

    # Metadata Labels and values
    ws["A19"] = "Executed At:"
    ws["A19"].font = bold_body_font
    ws["B19"] = timestamp
    ws["B19"].font = body_font

    ws["A20"] = "Database:"
    ws["A20"].font = bold_body_font
    ws["B20"] = database
    ws["B20"].font = body_font

    if status == "completed":
        ws["A21"] = "Execution Time:"
        ws["A21"].font = bold_body_font
        ws["B21"] = f"{duration_ms:.2f} ms"
        ws["B21"].font = body_font

        ws["A22"] = "Rows Returned:"
        ws["A22"].font = bold_body_font
        ws["B22"] = len(rows)
        ws["B22"].font = body_font

        # Row 24: Results Output Header
        ws["A24"] = "Query Output"
        ws["A24"].font = section_header_font

        # Row 25: Table Headers
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=25, column=col_idx)
            cell.value = header
            cell.font = table_header_font
            cell.fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_gray

        # Row 26+: Data rows
        zebra_fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
        white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        
        left_alignment = Alignment(horizontal="left", vertical="center")
        right_alignment = Alignment(horizontal="right", vertical="center")
        center_alignment = Alignment(horizontal="center", vertical="center")

        if len(rows) > 0:
            for r_offset, row in enumerate(rows):
                r_idx = 26 + r_offset
                fill_to_use = zebra_fill if r_offset % 2 == 1 else white_fill
                
                for c_idx, val in enumerate(row, 1):
                    cell = ws.cell(row=r_idx, column=c_idx)
                    
                    if isinstance(val, Decimal):
                        cleaned_val = float(val)
                    else:
                        cleaned_val = val
                        
                    cell.value = cleaned_val
                    cell.font = table_data_font
                    cell.fill = fill_to_use
                    cell.border = thin_gray

                    # Alignments and Formatting
                    if isinstance(cleaned_val, (int, float)):
                        cell.alignment = right_alignment
                        if isinstance(cleaned_val, float):
                            cell.number_format = "#,##0.00"
                        else:
                            cell.number_format = "#,##0"
                    elif isinstance(cleaned_val, (datetime.datetime, datetime.date)):
                        cell.alignment = center_alignment
                        cell.number_format = "yyyy-mm-dd"
                    elif cleaned_val is None or cleaned_val == "":
                        cell.value = "NULL"
                        cell.alignment = left_alignment
                        cell.font = Font(name="Segoe UI", size=10, italic=True, color="858585")
                    else:
                        cell.alignment = left_alignment
        else:
            # Empty rows warning cell
            ws.merge_cells(start_row=26, start_column=1, end_row=26, end_column=max(len(headers), 2))
            cell = ws.cell(row=26, column=1)
            cell.value = "Zero rows returned from database."
            cell.font = Font(name="Segoe UI", size=10, italic=True, color="858585")
            cell.fill = white_fill
            cell.border = thin_gray

        # Freeze headers row
        ws.freeze_panes = "A26"

    else:
        # FAILED Query error warning layout
        ws["A21"] = "Status:"
        ws["A21"].font = bold_body_font
        ws["B21"] = "FAILED"
        ws["B21"].font = Font(name="Segoe UI", size=10, bold=True, color="C00000")

        ws["A22"] = "Reason:"
        ws["A22"].font = bold_body_font
        
        ws.merge_cells("B22:H24")
        reason_cell = ws["B22"]
        reason_cell.value = error_message
        reason_cell.font = error_msg_font
        reason_cell.alignment = Alignment(wrap_text=True, vertical="top")
        
        red_side = Side(style="thin", color="C00000")
        for r in range(22, 25):
            for c in range(2, 9):
                cell = ws.cell(row=r, column=c)
                left = red_side if c == 2 else None
                right = red_side if c == 8 else None
                top = red_side if r == 22 else None
                bottom = red_side if r == 24 else None
                cell.border = Border(left=left, right=right, top=top, bottom=bottom)

    # 4. Auto-adjust columns widths based strictly on results headers & rows (from row 25 down)
    if status == "completed" and len(headers) > 0:
        for col_idx in range(1, len(headers) + 1):
            col_letter = get_column_letter(col_idx)
            max_len = len(headers[col_idx - 1]) # start with length of header text
            
            for r_idx in range(26, ws.max_row + 1):
                cell_val = ws.cell(row=r_idx, column=col_idx).value
                if cell_val is not None:
                    if isinstance(cell_val, (datetime.datetime, datetime.date)):
                        val_str = cell_val.strftime("%Y-%m-%d")
                    elif isinstance(cell_val, float):
                        val_str = f"{cell_val:,.2f}"
                    else:
                        val_str = str(cell_val)
                    
                    for line in val_str.split("\n"):
                        if len(line) > max_len:
                            max_len = len(line)
                            
            ws.column_dimensions[col_letter].width = max(
                min(max_len + 4, 50), 12
            )
    else:
        # Default fallback sizing if no headers/rows
        for col_idx in range(1, 9):
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = 14


def save_workbook(wb: openpyxl.Workbook, file_path: Path) -> None:
    """Saves the workbook to the specified file path, creating parent directories if needed.

    Args:
        wb: openpyxl Workbook instance.
        file_path: Output file destination path.
    """
    logger.info(f"Saving workbook to '{file_path}'...")
    file_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(file_path))
    logger.info("Workbook saved successfully.")
